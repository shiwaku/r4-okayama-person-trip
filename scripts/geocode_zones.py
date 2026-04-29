"""
マスターデータのゾーンコードに座標を付加するスクリプト
1. ゾーンコード表から (市町村コード, 大字・町コード) → 住所名 を構築
2. 国土地理院APIでジオコーディング
3. 座標ルックアップCSVを出力
"""
import openpyxl
import pandas as pd
import urllib.request
import urllib.parse
import json
import time
import os
from collections import Counter, defaultdict

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_DIR = os.path.join(ROOT_DIR, "source")
DATA_DIR = os.path.join(ROOT_DIR, "data")
CACHE_FILE = os.path.join(DATA_DIR, "geocode_cache.json")
OUTPUT_FILE = os.path.join(DATA_DIR, "zone_coords.csv")

# ── Step 1: ゾーンコード表から住所名ルックアップを構築 ──────────────────
print("Step 1: ゾーンコード表を読み込み中...")
wb = openpyxl.load_workbook(os.path.join(SOURCE_DIR, "03_コード表.xlsx"), read_only=True)
ws = wb["ｿﾞｰﾝｺｰﾄﾞ表"]
rows = list(ws.iter_rows(min_row=9, values_only=True))  # 8行目がヘッダー
wb.close()

# (city_code, town_code) → [(pref_name, city_name, town_name), ...]
zone_names = defaultdict(list)
for r in rows:
    pref = r[1]   # 都道府県名
    city = r[2]   # 市区町村名
    town = r[3]   # 大字町丁目名
    f    = r[5]   # 市区町村コード(5桁)
    g    = r[6]   # 大字・町コード(3桁)
    h    = r[7]   # 丁目コード(2桁)
    if not (pref and city and town and f and g is not None):
        continue
    try:
        city_code = str(f)
        town_code = str(int(str(g).lstrip("0") or "0"))  # "001" → "1"
    except (ValueError, TypeError):
        continue
    zone_names[(city_code, town_code)].append((pref, city, town, str(h or "00")))

print(f"  ゾーン種類数（コード表）: {len(zone_names)}")

# 各 (city_code, town_code) の代表住所を決定
# 優先: 丁目コード="00"（丁目なし） → なければ最初のエントリ
def pick_address(entries):
    for pref, city, town, chome in entries:
        if chome == "00":
            return pref, city, town
    return entries[0][0], entries[0][1], entries[0][2]

zone_address = {k: pick_address(v) for k, v in zone_names.items()}

# ── Step 2: マスターデータ（平日・休日）で使われているゾーンコードを収集 ──
print("Step 2: マスターデータのゾーンコードを収集中...")
all_codes = set()
for fname in ["01_R4岡山PTマスターデータ平日.xlsx", "02_R4岡山PTマスターデータ休日.xlsx"]:
    wb = openpyxl.load_workbook(os.path.join(SOURCE_DIR, fname), read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=11, values_only=True):
        if row[2] is None:
            continue
        for col_city, col_town in [(13, 14), (16, 17)]:  # 出発地, 目的地
            try:
                if row[col_city] and row[col_town] and str(row[col_town]) not in ("*", "None"):
                    all_codes.add((str(row[col_city]), str(int(float(row[col_town])))))
            except (ValueError, TypeError):
                pass
    wb.close()

print(f"  ユニークゾーン数: {len(all_codes)}")

# ── Step 3: ジオコーディング（キャッシュ付き） ────────────────────────
print("Step 3: ジオコーディング中...")

# キャッシュ読み込み
cache = {}
if os.path.exists(CACHE_FILE):
    with open(CACHE_FILE, encoding="utf-8") as f:
        cache = json.load(f)
    print(f"  キャッシュ読み込み: {len(cache)}件")

def geocode(address):
    """国土地理院APIで住所→(lat,lon)"""
    if address in cache:
        return cache[address]
    url = f"https://msearch.gsi.go.jp/address-search/AddressSearch?q={urllib.parse.quote(address)}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read())
        if data:
            coords = data[0]["geometry"]["coordinates"]
            result = {"lon": coords[0], "lat": coords[1]}
        else:
            result = {"lon": None, "lat": None}
    except Exception as e:
        result = {"lon": None, "lat": None}
    cache[address] = result
    time.sleep(0.3)
    return result

# 未キャッシュのアドレスを構築してジオコーディング
to_geocode = {}
for code in all_codes:
    if code in zone_address:
        pref, city, town = zone_address[code]
        addr = pref + city + town
    else:
        # コード表に存在しない（県外など）: 市町村名のみ
        addr = None
    to_geocode[code] = addr

unique_addresses = set(a for a in to_geocode.values() if a and a not in cache)
total = len(unique_addresses)
print(f"  未キャッシュのアドレス: {total}件")

done = 0
for addr in unique_addresses:
    geocode(addr)
    done += 1
    if done % 100 == 0:
        # 途中経過とキャッシュ保存
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False)
        print(f"  進捗: {done}/{total}")

# 最終保存
with open(CACHE_FILE, "w", encoding="utf-8") as f:
    json.dump(cache, f, ensure_ascii=False)
print(f"  完了: {done}件ジオコーディング, キャッシュ保存済み")

# ── Step 4: ゾーン座標テーブルCSV出力 ──────────────────────────────
print("Step 4: 座標テーブルを出力中...")

records = []
no_address = 0
no_coords = 0

for code in sorted(all_codes):
    city_code, town_code = code
    addr = to_geocode.get(code)
    if addr:
        result = cache.get(addr, {"lon": None, "lat": None})
    else:
        result = {"lon": None, "lat": None}
        no_address += 1

    if result["lat"] is None:
        no_coords += 1

    pref_name, city_name, town_name = zone_address.get(code, ("", "", ""))
    records.append({
        "city_code": city_code,
        "town_code": town_code,
        "pref_name": pref_name,
        "city_name": city_name,
        "town_name": town_name,
        "address": addr or "",
        "lat": result["lat"],
        "lon": result["lon"],
    })

df = pd.DataFrame(records)
df.to_csv(OUTPUT_FILE, index=False, encoding="utf-8-sig")
print(f"  出力: {OUTPUT_FILE}")
print(f"  総ゾーン数: {len(records)}")
print(f"  住所名なし: {no_address}件")
print(f"  座標取得失敗: {no_coords}件")

# ── Step 5: GeoJSON出力 ─────────────────────────────────────────────
print("Step 5: GeoJSONを出力中...")
GEOJSON_FILE = os.path.join(DATA_DIR, "zone_coords.geojson")

features = []
for rec in records:
    if rec["lat"] is None or rec["lon"] is None:
        continue
    features.append({
        "type": "Feature",
        "geometry": {
            "type": "Point",
            "coordinates": [rec["lon"], rec["lat"]],
        },
        "properties": {
            "city_code": rec["city_code"],
            "town_code": rec["town_code"],
            "pref_name": rec["pref_name"],
            "city_name": rec["city_name"],
            "town_name": rec["town_name"],
            "address":   rec["address"],
        },
    })

geojson = {"type": "FeatureCollection", "features": features}
with open(GEOJSON_FILE, "w", encoding="utf-8") as f:
    json.dump(geojson, f, ensure_ascii=False, indent=2)
print(f"  出力: {GEOJSON_FILE} ({len(features)}点)")
print("Done.")
