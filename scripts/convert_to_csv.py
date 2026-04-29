"""
Excel → CSV 一括変換スクリプト

出力先: data/csv/
  01_master_weekday.csv      平日マスターデータ（行11〜）
  02_master_holiday.csv      休日マスターデータ（行11〜）
  code_調査日ロット番号.csv
  code_性別.csv
  code_就業就学状況.csv
  code_自動車免許.csv
  code_自由に使える自動車.csv
  code_移動の有無.csv
  code_施設の種類.csv
  code_移動目的.csv
  code_交通手段.csv
  code_代表交通手段.csv
  code_年齢階層.csv
  code_時間帯コード.csv
  code_午前午後.csv
  code_ゾーンコード.csv
  code_ターミナルコード.csv
"""

import pandas as pd
import openpyxl
from pathlib import Path

ROOT_DIR = Path(__file__).parent.parent
SOURCE_DIR = ROOT_DIR / "source"
DATA_DIR = ROOT_DIR / "data"
CSV_DIR = DATA_DIR / "csv"
CSV_DIR.mkdir(exist_ok=True)


# ── マスターデータ用ヘッダー生成（Excel行6〜9から構築） ──────────────────────

SKIP_PHRASES = {
    "問1：あなたご自身のことについて",
    "問2：調査日にどこかへ移動しましたか？",
    "移動状況", "集計用コード",
    "※ここからは、県が作業用に追加した項目です",
    "●：当該トリップにおいて、1度でも利用された交通手段",
    "当該トリップにおいて、各交通手段が使われた延べ回数（実数）",
    "当該トリップにおいて、各交通手段が使われた延べ回数（拡大処理後）",
    "*",
}
GROUP_PREFIX = {
    6: "自宅住所_", 7: "自宅住所_",
    13: "出発地_", 14: "出発地_", 15: "出発地_",
    16: "目的地_", 17: "目的地_", 18: "目的地_",
    19: "出発_", 20: "出発_", 21: "出発_",
    22: "到着_", 23: "到着_", 24: "到着_",
    26: "交通手段_", 27: "交通手段_", 28: "交通手段_", 29: "交通手段_", 30: "交通手段_",
    31: "利用ターミナル_", 32: "利用ターミナル_", 33: "利用ターミナル_",
    34: "利用ターミナル_", 35: "利用ターミナル_", 36: "利用ターミナル_",
    48: "フラグ_", 49: "フラグ_", 50: "フラグ_", 51: "フラグ_", 52: "フラグ_",
    53: "フラグ_", 54: "フラグ_", 55: "フラグ_", 56: "フラグ_", 57: "フラグ_",
    58: "フラグ_", 59: "フラグ_", 60: "フラグ_", 61: "フラグ_", 62: "フラグ_",
    65: "延べ_", 66: "延べ_", 67: "延べ_", 68: "延べ_", 69: "延べ_",
    70: "延べ_", 71: "延べ_", 72: "延べ_", 73: "延べ_", 74: "延べ_",
    75: "延べ_", 76: "延べ_", 77: "延べ_", 78: "延べ_",
    80: "拡大_", 81: "拡大_", 82: "拡大_", 83: "拡大_", 84: "拡大_",
    85: "拡大_", 86: "拡大_", 87: "拡大_", 88: "拡大_", 89: "拡大_",
    90: "拡大_", 91: "拡大_", 92: "拡大_", 93: "拡大_",
}
OVERRIDE_NAMES = {
    2:  "SEQ", 3:  "サンプルNO", 4:  "調査日_ロット番号",
    37: "移動回数", 38: "拡大係数", 39: "年齢階層コード",
    40: "代表交通手段", 41: "時間帯コード",
    44: "発地_県内フラグ", 45: "着地_県内フラグ", 47: "手段数",
    53: "フラグ_自動車計", 64: "手段数_延べ", 79: "手段数_拡大後",
}

def build_col_names() -> list[str]:
    wb = openpyxl.load_workbook(
        SOURCE_DIR / "01_R4岡山PTマスターデータ平日.xlsx", read_only=True, data_only=True
    )
    ws = wb["平日"]
    header_rows = [list(r) for r in ws.iter_rows(min_row=6, max_row=9, values_only=True)]
    wb.close()

    col_names = []
    for col in range(94):
        if col in OVERRIDE_NAMES:
            col_names.append(OVERRIDE_NAMES[col])
            continue
        parts = []
        for row in header_rows:
            v = row[col] if col < len(row) else None
            if v is not None:
                s = str(v).replace("\n", " ").strip()
                if s and s not in SKIP_PHRASES:
                    parts.append(s)
        label = "_".join(parts) if parts else f"col{col}"
        prefix = GROUP_PREFIX.get(col, "")
        if prefix:
            bare = prefix.rstrip("_")
            if not label.startswith(bare):
                label = prefix + label
        col_names.append(label)

    # 重複排除
    seen: dict = {}
    unique = []
    for name in col_names:
        if name in seen:
            seen[name] += 1
            unique.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            unique.append(name)

    # A列・B列（index 0,1）を除去して返す
    return unique[2:]


# ── マスターデータ ──────────────────────────────────────────────────────────

def convert_master(fname: str, out_name: str, col_names: list[str]):
    path = SOURCE_DIR / fname
    print(f"{fname} → {out_name} ...")
    df = pd.read_excel(path, header=None, skiprows=10, dtype=str)
    df = df.drop(columns=[0, 1])
    df.columns = col_names
    df.to_csv(CSV_DIR / out_name, index=False, encoding="utf-8-sig")
    print(f"  {len(df)} 行, {len(df.columns)} 列")


# ── コード表: 指定セル範囲を抜き出してCSV化 ─────────────────────────────────

def extract_table(ws, title: str, out_name: str, header_row: int, data_row_start: int,
                  data_row_end: int, col_start: int, col_end: int):
    """
    ws: openpyxl worksheet
    header_row, data_row_*: 1-indexed行番号
    col_start, col_end: 1-indexed列番号
    """
    def row_vals(r):
        return [ws.cell(row=r, column=c).value for c in range(col_start, col_end + 1)]

    headers = row_vals(header_row)
    rows = [row_vals(r) for r in range(data_row_start, data_row_end + 1)
            if any(v is not None for v in row_vals(r))]

    df = pd.DataFrame(rows, columns=headers)
    df.to_csv(CSV_DIR / out_name, index=False, encoding="utf-8-sig")
    print(f"  {title}: {len(df)} 行 → {out_name}")


def convert_code_tables():
    path = SOURCE_DIR / "03_コード表.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # ── コード表シート ──────────────────────────────────────────────────────
    ws = wb["コード表"]
    print("コード表シート:")

    # A-B列グループ
    extract_table(ws, "調査日ロット番号",    "code_調査日ロット番号.csv",    4,  5,  6,  1, 2)
    extract_table(ws, "性別",               "code_性別.csv",               9, 10, 11,  1, 2)
    extract_table(ws, "就業就学状況",        "code_就業就学状況.csv",       14, 15, 22,  1, 2)
    extract_table(ws, "自動車免許",          "code_自動車免許.csv",         25, 26, 30,  1, 2)
    extract_table(ws, "自由に使える自動車",  "code_自由に使える自動車.csv", 33, 34, 36,  1, 2)
    extract_table(ws, "移動の有無",          "code_移動の有無.csv",         39, 40, 41,  1, 2)
    extract_table(ws, "交通手段",            "code_交通手段.csv",           44, 45, 58,  1, 2)

    # D-E列グループ
    extract_table(ws, "施設の種類",  "code_施設の種類.csv",  4,  5, 19,  4, 5)
    extract_table(ws, "移動目的",    "code_移動目的.csv",   22, 23, 31,  4, 5)
    extract_table(ws, "時間帯コード","code_時間帯コード.csv",34, 35, 39,  4, 5)

    # D-F列（代表交通手段は3列）
    extract_table(ws, "代表交通手段","code_代表交通手段.csv",44, 45, 58,  4, 6)

    # G-H列グループ
    extract_table(ws, "年齢階層",   "code_年齢階層.csv",    4,  5, 22,  7, 8)
    extract_table(ws, "午前午後",   "code_午前午後.csv",   35, 36, 39,  7, 8)

    # ── ゾーンコード表シート ───────────────────────────────────────────────
    ws2 = wb["ｿﾞｰﾝｺｰﾄﾞ表"]
    print("ゾーンコード表シート:")

    # 列数を確認
    max_col = 0
    for row in ws2.iter_rows(min_row=8, max_row=8, values_only=False):
        for cell in row:
            if cell.value is not None:
                max_col = max(max_col, cell.column)

    headers = [ws2.cell(row=8, column=c).value for c in range(1, max_col + 1)]
    rows = []
    for r in ws2.iter_rows(min_row=10, values_only=True):
        vals = list(r[:max_col])
        if any(v is not None for v in vals):
            rows.append(vals)

    df_zone = pd.DataFrame(rows, columns=headers)
    df_zone.to_csv(CSV_DIR / "code_ゾーンコード.csv", index=False, encoding="utf-8-sig")
    print(f"  ゾーンコード: {len(df_zone)} 行 → code_ゾーンコード.csv")

    # ── ターミナルコード表シート ───────────────────────────────────────────
    ws3 = wb["ターミナルコード表"]
    print("ターミナルコード表シート:")

    # C列〜H列（コード4桁〜備考）
    headers_t = [ws3.cell(row=2, column=c).value for c in range(3, 9)]
    rows_t = []
    for r in ws3.iter_rows(min_row=3, values_only=True):
        vals = list(r[2:8])
        if any(v is not None for v in vals):
            rows_t.append(vals)

    df_term = pd.DataFrame(rows_t, columns=headers_t)
    df_term.to_csv(CSV_DIR / "code_ターミナルコード.csv", index=False, encoding="utf-8-sig")
    print(f"  ターミナルコード: {len(df_term)} 行 → code_ターミナルコード.csv")

    wb.close()


if __name__ == "__main__":
    print("=== マスターデータ変換 ===")
    col_names = build_col_names()
    convert_master("01_R4岡山PTマスターデータ平日.xlsx", "01_master_weekday.csv", col_names)
    convert_master("02_R4岡山PTマスターデータ休日.xlsx", "02_master_holiday.csv", col_names)

    print("\n=== コード表変換 ===")
    convert_code_tables()

    print(f"\n完了: {CSV_DIR}")
