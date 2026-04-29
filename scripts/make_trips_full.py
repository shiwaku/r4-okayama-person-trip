"""
マスターデータ（平日・休日）の全94列に座標を付加し、コード値を読み替えてCSV出力する。

- ヘッダーはExcel行6〜9から自動生成（Excelの列名を忠実に再現）
- コード値は日本語ラベルに変換
- 出発地・目的地の緯度経度を末尾に追加

出力: trips_full.csv
"""
import os
import pandas as pd
import openpyxl

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_DIR = os.path.join(ROOT_DIR, "source")
DATA_DIR = os.path.join(ROOT_DIR, "data")

# ── コード定義 ────────────────────────────────────────────────────────────
GENDER      = {1: "男性", 2: "女性"}
SURVEY_DATE = {1: "平日10/12・休日10/16", 2: "平日10/19・休日10/23"}
EMPLOYMENT  = {
    1: "会社員等", 2: "自営(自宅)", 3: "自営(自宅外)",
    4: "農林水産", 5: "主婦(夫)", 6: "生徒・学生", 7: "無職", 9: "未記入",
}
LICENSE  = {1: "自動車", 2: "二輪・原付", 3: "自主返納", 4: "なし", 9: "未記入"}
HAS_CAR  = {1: "有", 2: "無", 9: "未記入"}
HAS_TRIP = {1: "有", 2: "無"}
FACILITY = {
    1: "住宅・寮", 2: "学校等", 3: "文化・宗教", 4: "医療・福祉",
    5: "事務所・会社", 6: "官公庁", 7: "問屋・卸売", 8: "商業施設",
    9: "飲食店", 10: "宿泊・娯楽", 11: "工場・倉庫", 12: "交通・運輸",
    13: "農林漁業", 14: "その他", 99: "未記入",
}
PURPOSE  = {
    1: "出勤", 2: "登校", 3: "業務", 4: "買物",
    5: "通院", 6: "私用", 7: "帰宅", 8: "その他", 9: "未記入",
}
MODE = {
    1: "徒歩", 2: "自転車", 3: "バイク",
    4: "自動車(運転)", 5: "自動車(同乗)",
    6: "路線バス", 7: "貸切バス", 8: "鉄道",
    9: "路面電車", 10: "タクシー", 11: "デマンドタクシー",
    12: "船舶・飛行機", 13: "その他", 99: "未記入",
}
AGE_GROUP = {
    4: "15-17歳", 5: "18-19歳", 6: "20-24歳", 7: "25-29歳",
    8: "30-34歳", 9: "35-39歳", 10: "40-44歳", 11: "45-49歳",
    12: "50-54歳", 13: "55-59歳", 14: "60-64歳", 15: "65-69歳",
    16: "70-74歳", 17: "75-79歳", 18: "80-84歳", 19: "85-89歳",
    20: "90-94歳", 21: "95歳-",
}
TIME_ZONE = {
    1: "朝ピーク(7-9時)", 2: "オフピーク(9-16時)",
    3: "夕ピーク(16-19時)", 4: "その他(-7時・19時-)", 9: "不明",
}
AM_PM = {1: "午前", 2: "午後"}

# 列インデックス → 変換マッピング（県内フラグは関数で処理）
CODE_MAP = {
    4:  SURVEY_DATE,
    5:  GENDER,
    8:  EMPLOYMENT,
    9:  LICENSE,
    10: HAS_CAR,
    11: HAS_TRIP,
    15: FACILITY,
    18: FACILITY,
    19: AM_PM,
    22: AM_PM,
    25: PURPOSE,
    26: MODE, 27: MODE, 28: MODE, 29: MODE, 30: MODE,
    39: AGE_GROUP,
    40: MODE,
    41: TIME_ZONE,
}


def cv(mapping, val):
    if val is None:
        return None
    try:
        return mapping.get(int(val), val)
    except (ValueError, TypeError):
        return val


def safe_code(v):
    if v is None or str(v) in ("*", "None", ""):
        return None
    try:
        return str(int(float(v)))
    except (ValueError, TypeError):
        return None


def in_pref_label(v):
    """●→県内、None/空→県外"""
    if v == "●":
        return "県内"
    if v is None or str(v).strip() == "":
        return "県外"
    return v


# ── ヘッダー生成（Excel行6〜9から自動構築） ─────────────────────────────
print("ヘッダー読み込み中...")
wb_hdr = openpyxl.load_workbook(
    os.path.join(SOURCE_DIR, "01_R4岡山PTマスターデータ平日.xlsx"), read_only=True
)
ws_hdr = wb_hdr["平日"]
header_rows = [list(r) for r in ws_hdr.iter_rows(min_row=6, max_row=9, values_only=True)]
wb_hdr.close()

# 省略する長い説明フレーズ
SKIP_PHRASES = {
    "問1：あなたご自身のことについて",
    "問2：調査日にどこかへ移動しましたか？",
    "移動状況",
    "集計用コード",
    "※ここからは、県が作業用に追加した項目です",
    "●：当該トリップにおいて、1度でも利用された交通手段",
    "当該トリップにおいて、各交通手段が使われた延べ回数（実数）",
    "当該トリップにおいて、各交通手段が使われた延べ回数（拡大処理後）",
    "*",
}

# 列グループのコンテキストプレフィックス（上位行の文脈を保持）
GROUP_PREFIX = {
    6: "自宅住所_", 7: "自宅住所_",
    13: "出発地_", 14: "出発地_", 15: "出発地_",
    16: "目的地_", 17: "目的地_", 18: "目的地_",
    19: "出発_", 20: "出発_", 21: "出発_",
    22: "到着_", 23: "到着_", 24: "到着_",
    26: "交通手段_", 27: "交通手段_", 28: "交通手段_", 29: "交通手段_", 30: "交通手段_",
    31: "利用ターミナル_", 32: "利用ターミナル_", 33: "利用ターミナル_",
    34: "利用ターミナル_", 35: "利用ターミナル_", 36: "利用ターミナル_",
    48: "フラグ_", 49: "フラグ_", 50: "フラグ_", 51: "フラグ_", 52: "フラグ_",
    53: "フラグ_", 54: "フラグ_", 55: "フラグ_", 56: "フラグ_", 57: "フラグ_",
    58: "フラグ_", 59: "フラグ_", 60: "フラグ_", 61: "フラグ_", 62: "フラグ_",
    65: "延べ_", 66: "延べ_", 67: "延べ_", 68: "延べ_", 69: "延べ_",
    70: "延べ_", 71: "延べ_", 72: "延べ_", 73: "延べ_", 74: "延べ_",
    75: "延べ_", 76: "延べ_", 77: "延べ_", 78: "延べ_",
    80: "拡大_", 81: "拡大_", 82: "拡大_", 83: "拡大_", 84: "拡大_",
    85: "拡大_", 86: "拡大_", 87: "拡大_", 88: "拡大_", 89: "拡大_",
    90: "拡大_", 91: "拡大_", 92: "拡大_", 93: "拡大_",
}

# 曖昧になりがちな列を明示的に上書き
OVERRIDE_NAMES = {
    2:  "SEQ",
    3:  "サンプルNO",
    4:  "調査日_ロット番号",
    37: "移動回数",
    38: "拡大係数",
    39: "年齢階層コード",
    40: "代表交通手段",
    41: "時間帯コード",
    44: "発地_県内フラグ",
    45: "着地_県内フラグ",
    47: "手段数",
    53: "フラグ_自動車計",
    64: "手段数_延べ",
    79: "手段数_拡大後",
}

NUM_COLS = 94
col_names = []
for col in range(NUM_COLS):
    if col in OVERRIDE_NAMES:
        col_names.append(OVERRIDE_NAMES[col])
        continue
    parts = []
    for row in header_rows:
        v = row[col] if col < len(row) else None
        if v is not None:
            s = str(v).replace("\n", " ").strip()
            if s and s not in SKIP_PHRASES:
                parts.append(s)
    label = "_".join(parts) if parts else f"col{col}"
    prefix = GROUP_PREFIX.get(col, "")
    if prefix:
        # プレフィックスがすでに含まれていなければ先頭に付ける
        bare = prefix.rstrip("_")
        if not label.startswith(bare):
            label = prefix + label
    col_names.append(label)

# 重複排除（同名列に _2, _3 ... を付与）
seen: dict = {}
unique_col_names = []
for name in col_names:
    if name in seen:
        seen[name] += 1
        unique_col_names.append(f"{name}_{seen[name]}")
    else:
        seen[name] = 0
        unique_col_names.append(name)

print("  生成されたヘッダー:")
for i, n in enumerate(unique_col_names):
    print(f"    col{i:2d}: {n}")


# ── 座標ルックアップ ─────────────────────────────────────────────────────
print("\n座標ルックアップ読み込み中...")
coords_df = pd.read_csv(
    os.path.join(DATA_DIR, "zone_coords.csv"),
    dtype={"city_code": str, "town_code": str},
    encoding="utf-8",
)
coords = {
    (r.city_code, r.town_code): (r.lat, r.lon)
    for r in coords_df.itertuples()
}
print(f"  {len(coords)} ゾーン")


# ── ターミナルコード（列2=コード, 列3=名称） ─────────────────────────────
print("ターミナルコード読み込み中...")
wb_code = openpyxl.load_workbook(
    os.path.join(SOURCE_DIR, "03_コード表.xlsx"), read_only=True
)
terminal = {}
for row in wb_code["ターミナルコード表"].iter_rows(min_row=3, values_only=True):
    code = row[2]
    name = row[3]
    if code is not None:
        try:
            terminal[int(code)] = str(name) if name else ""
        except (ValueError, TypeError):
            pass
wb_code.close()
print(f"  {len(terminal)} ターミナル")


# ── データ処理 ───────────────────────────────────────────────────────────
def read_and_process(path, day_type):
    print(f"\n{day_type} 読み込み中: {os.path.basename(path)}")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    result = []
    miss_o = miss_d = 0

    for row in ws.iter_rows(min_row=11, values_only=True):
        if row[2] is None:
            continue
        row = list(row)

        # ターミナルコード → 名称
        for ti in range(31, 37):
            v = row[ti]
            if v is not None:
                try:
                    row[ti] = terminal.get(int(v), v)
                except (ValueError, TypeError):
                    pass

        # コード値 → ラベル
        for col_idx, mapping in CODE_MAP.items():
            row[col_idx] = cv(mapping, row[col_idx])

        # 県内フラグ（● = 県内）
        row[44] = in_pref_label(row[44])
        row[45] = in_pref_label(row[45])

        # 座標結合
        o_city = safe_code(row[13])
        o_town = safe_code(row[14])
        d_city = safe_code(row[16])
        d_town = safe_code(row[17])
        o_lat, o_lon = coords.get((o_city, o_town), (None, None))
        d_lat, d_lon = coords.get((d_city, d_town), (None, None))
        if o_lat is None: miss_o += 1
        if d_lat is None: miss_d += 1

        rec = {"平休区分": day_type}
        for i, name in enumerate(unique_col_names):
            rec[name] = row[i] if i < len(row) else None
        rec["出発地緯度"] = o_lat
        rec["出発地経度"] = o_lon
        rec["目的地緯度"] = d_lat
        rec["目的地経度"] = d_lon
        result.append(rec)

    wb.close()
    print(f"  {len(result):,} 行  座標なし: 出発地 {miss_o}件, 目的地 {miss_d}件")
    return result


all_records = []
for fname, day_type in [
    ("01_R4岡山PTマスターデータ平日.xlsx", "平日"),
    ("02_R4岡山PTマスターデータ休日.xlsx", "休日"),
]:
    all_records.extend(read_and_process(os.path.join(SOURCE_DIR, fname), day_type))

print("\nCSV出力中...")
df = pd.DataFrame(all_records)
out_path = os.path.join(DATA_DIR, "trips_full.csv")
df.to_csv(out_path, index=False, encoding="utf-8")
print(f"出力完了: {out_path}")
print(f"総トリップ数: {len(df):,} 行  列数: {len(df.columns)}")
print("\n先頭1行（主要列）:")
show_cols = [
    "平休区分", "SEQ", "サンプルNO", "調査日_ロット番号", "性別", "年齢階層コード",
    "就業・就学状況", "移動の有無", "トリップ番号", "移動目的", "代表交通手段",
    "時間帯コード", "拡大係数", "発地_県内フラグ", "着地_県内フラグ",
    "出発地緯度", "出発地経度", "目的地緯度", "目的地経度",
]
print(df[show_cols].head(3).to_string())
